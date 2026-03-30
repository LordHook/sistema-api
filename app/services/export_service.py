"""Servicio de exportación a Excel y PDF."""
import io
import calendar
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib import colors as rl_colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm

from app.services.schedule_generator import get_schedule_grid
from app.models.audit import AuditLog
from app.models.worker import Worker
from app.models.attendance import AttendanceRecord
from sqlalchemy import extract


# Color mapping for Excel
SHIFT_FILLS = {
    'M': PatternFill(start_color='22c55e', end_color='22c55e', fill_type='solid'),
    'T': PatternFill(start_color='f59e0b', end_color='f59e0b', fill_type='solid'),
    'N': PatternFill(start_color='6366f1', end_color='6366f1', fill_type='solid'),
    'D': PatternFill(start_color='94a3b8', end_color='94a3b8', fill_type='solid'),
    'V': PatternFill(start_color='06b6d4', end_color='06b6d4', fill_type='solid'),
    'C': PatternFill(start_color='a855f7', end_color='a855f7', fill_type='solid'),
    'R': PatternFill(start_color='ef4444', end_color='ef4444', fill_type='solid'),
}

SHIFT_PDF_COLORS = {
    'M': rl_colors.HexColor('#22c55e'),
    'T': rl_colors.HexColor('#f59e0b'),
    'N': rl_colors.HexColor('#6366f1'),
    'D': rl_colors.HexColor('#94a3b8'),
    'V': rl_colors.HexColor('#06b6d4'),
    'C': rl_colors.HexColor('#a855f7'),
    'R': rl_colors.HexColor('#ef4444'),
}


def export_schedule_excel(year, month):
    """Export the monthly schedule to an Excel file."""
    grid = get_schedule_grid(year, month)
    wb = Workbook()
    ws = wb.active
    ws.title = f'Rol {grid["month_name"]} {year}'

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    title_font = Font(bold=True, size=14, color='1e3a5f')

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=grid['num_days'] + 4)
    title_cell = ws['A1']
    title_cell.value = f'ROL DE SERVICIO - CCO - {grid["month_name"].upper()} {year}'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    current_row = 3

    # Column headers
    headers = ['N°', 'R.L', 'APELLIDOS Y NOMBRES']
    for dh in grid['day_headers']:
        headers.append(f'{dh["weekday"]}\n{dh["day"]}')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 28
    for col in range(4, grid['num_days'] + 4):
        ws.column_dimensions[get_column_letter(col)].width = 4.5

    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # Sections
    for section in grid['sections']:
        # Section header
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=grid['num_days'] + 3
        )
        sec_cell = ws.cell(row=current_row, column=1,
                           value=f'SECCIÓN {section["key"]}: {section["name"]}')
        sec_cell.font = Font(bold=True, size=11, color='FFFFFF')
        sec_cell.fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')
        sec_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for group in section['groups']:
            if 'label' in group:
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row, end_column=grid['num_days'] + 3
                )
                grp_cell = ws.cell(row=current_row, column=1, value=f'  {group["label"]}')
                grp_cell.font = Font(bold=True, size=10, color='94a3b8')
                grp_cell.fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
                ws.row_dimensions[current_row].height = 20
                current_row += 1

            rows = group.get('rows', [])
            if isinstance(group, dict) and 'rows' not in group and len(group) > 0:
                rows = group.get('rows', [])

            for row in rows:
                worker = row['worker']
                ws.cell(row=current_row, column=1, value=worker['order_number']).border = thin_border
                ws.cell(row=current_row, column=2, value=worker['regime']).border = thin_border
                name_cell = ws.cell(row=current_row, column=3, value=worker['name'])
                name_cell.border = thin_border
                name_cell.alignment = Alignment(horizontal='left')

                for day_data in row['days']:
                    col = day_data['day'] + 3
                    cell = ws.cell(row=current_row, column=col, value=day_data['shift'])
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    cell.font = Font(bold=True, size=9, color='FFFFFF')
                    if day_data['shift'] in SHIFT_FILLS:
                        cell.fill = SHIFT_FILLS[day_data['shift']]

                ws.row_dimensions[current_row].height = 18
                current_row += 1

        current_row += 1  # Space between sections

    # Legend
    current_row += 1
    ws.cell(row=current_row, column=1, value='LEYENDA:').font = Font(bold=True, size=10)
    current_row += 1
    legends = [
        ('M', 'Mañana (06:00-14:00)'),
        ('T', 'Tarde (14:00-22:00)'),
        ('N', 'Noche (22:00-06:00)'),
        ('D', 'Descanso'),
        ('V', 'Vacaciones'),
        ('C', 'Compensado'),
        ('R', 'Renuncia'),
    ]
    for code, desc in legends:
        code_cell = ws.cell(row=current_row, column=1, value=code)
        code_cell.fill = SHIFT_FILLS.get(code)
        code_cell.font = Font(bold=True, color='FFFFFF')
        code_cell.alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=2, value=desc).font = Font(size=9)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
        current_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_attendance_excel(year, month, group_filter=None, user_role='admin'):
    """Export the monthly attendance to an Excel file."""
    grid = get_schedule_grid(year, month, group_filter, user_role)
    
    # Fetch all attendance records for this month
    records = AttendanceRecord.query.filter(
        extract('year', AttendanceRecord.attendance_date) == year,
        extract('month', AttendanceRecord.attendance_date) == month
    ).all()
    att_map = {(r.worker_id, r.attendance_date.day): r for r in records}

    wb = Workbook()
    ws = wb.active
    ws.title = f'Asist {grid["month_name"]} {year}'

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    header_fill = PatternFill(start_color='0f766e', end_color='0f766e', fill_type='solid') # Teal base
    header_font = Font(bold=True, color='FFFFFF', size=10)
    title_font = Font(bold=True, size=14, color='0f766e')
    
    # Status colors
    att_colors = {
        'A': PatternFill(start_color='22c55e', end_color='22c55e', fill_type='solid'),
        'F': PatternFill(start_color='ef4444', end_color='ef4444', fill_type='solid'),
        'asistio': PatternFill(start_color='22c55e', end_color='22c55e', fill_type='solid'),
        'falto': PatternFill(start_color='ef4444', end_color='ef4444', fill_type='solid'),
        'tardanza': PatternFill(start_color='f59e0b', end_color='f59e0b', fill_type='solid'),
        'PO': PatternFill(start_color='6366f1', end_color='6366f1', fill_type='solid'),
        'PC': PatternFill(start_color='a855f7', end_color='a855f7', fill_type='solid'),
        'PV': PatternFill(start_color='ec4899', end_color='ec4899', fill_type='solid'),
        'DM': PatternFill(start_color='14b8a6', end_color='14b8a6', fill_type='solid'),
        'V': PatternFill(start_color='84cc16', end_color='84cc16', fill_type='solid'),
        'LM': PatternFill(start_color='d946ef', end_color='d946ef', fill_type='solid'),
        'LE': PatternFill(start_color='a3a3a3', end_color='a3a3a3', fill_type='solid'),
        'PS': PatternFill(start_color='f97316', end_color='f97316', fill_type='solid'),
        'NI': PatternFill(start_color='64748b', end_color='64748b', fill_type='solid'),
    }

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=grid['num_days'] + 4)
    title_cell = ws['A1']
    title_cell.value = f'CONTROL DE ASISTENCIA - CCO - {grid["month_name"].upper()} {year}'
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    current_row = 3
    headers = ['N°', 'R.L', 'APELLIDOS Y NOMBRES']
    for dh in grid['day_headers']:
        headers.append(f'{dh["weekday"]}\n{dh["day"]}')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 28
    for col in range(4, grid['num_days'] + 4):
        ws.column_dimensions[get_column_letter(col)].width = 4.5

    current_row += 1

    for section in grid['sections']:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=grid['num_days'] + 3)
        sec_cell = ws.cell(row=current_row, column=1, value=f'SECCIÓN {section["key"]}: {section["name"]}')
        sec_cell.font = Font(bold=True, size=11, color='FFFFFF')
        sec_cell.fill = PatternFill(start_color='115e59', end_color='115e59', fill_type='solid') # Dark Teal
        current_row += 1

        for group in section['groups']:
            if 'label' in group:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=grid['num_days'] + 3)
                grp_cell = ws.cell(row=current_row, column=1, value=f'  {group["label"]}')
                grp_cell.font = Font(bold=True, size=10, color='e2e8f0')
                grp_cell.fill = PatternFill(start_color='0f766e', end_color='0f766e', fill_type='solid')
                current_row += 1

            rows_data = group.get('rows', [])
            for row in rows_data:
                worker = row['worker']
                ws.cell(row=current_row, column=1, value=worker['order_number']).border = thin_border
                ws.cell(row=current_row, column=2, value=worker['regime']).border = thin_border
                ws.cell(row=current_row, column=3, value=worker['name']).border = thin_border

                for day_data in row['days']:
                    col = day_data['day'] + 3
                    record = att_map.get((worker['id'], day_data['day']))
                    att_status = record.status if record else ''
                    
                    # Convert to visual format
                    display_text = ''
                    if att_status == 'A': display_text = 'A'
                    elif att_status == 'F': display_text = 'F'
                    elif att_status == 'tardanza': display_text = 'T'
                    elif att_status == 'asistio': display_text = 'A'
                    elif att_status == 'falto': display_text = 'F'
                    elif att_status: display_text = att_status
                    elif day_data.get('shift') and day_data.get('shift') not in ['M', 'T', 'N', 'CLEAR', 'NI']:
                        display_text = day_data.get('shift')
                    
                    cell = ws.cell(row=current_row, column=col, value=display_text)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    cell.font = Font(bold=True, size=8, color='FFFFFF')
                    
                    if att_status in att_colors:
                        cell.fill = att_colors[att_status]
                    elif att_status.startswith('T '):
                        cell.fill = att_colors['tardanza']
                    elif att_status in SHIFT_FILLS:
                        cell.fill = SHIFT_FILLS.get(att_status)
                    elif not att_status and day_data.get('shift') not in ['M', 'T', 'N', 'CLEAR', 'NI', None, '']:
                        shift_val = day_data.get('shift')
                        if shift_val in SHIFT_FILLS:
                            cell.fill = SHIFT_FILLS.get(shift_val)
                        elif shift_val in att_colors:
                            cell.fill = att_colors.get(shift_val)

                current_row += 1
        current_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_schedule_pdf(year, month):
    """Export the monthly schedule to PDF."""
    grid = get_schedule_grid(year, month)

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A3),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = Paragraph(
        f'<b>ROL DE SERVICIO - CCO - {grid["month_name"].upper()} {year}</b>',
        styles['Title']
    )
    elements.append(title)
    elements.append(Spacer(1, 5 * mm))

    for section in grid['sections']:
        elements.append(Paragraph(
            f'<b>SECCIÓN {section["key"]}: {section["name"]}</b>',
            styles['Heading3']
        ))

        # Build table data
        header_row = ['N°', 'R.L', 'Nombre']
        header_row += [str(dh['day']) for dh in grid['day_headers']]

        table_data = [header_row]

        for group in section['groups']:
            rows = group.get('rows', [])
            for row in rows:
                w = row['worker']
                data_row = [str(w['order_number']), w['regime'], w['name']]
                data_row += [d['shift'] for d in row['days']]
                table_data.append(data_row)

        if len(table_data) > 1:
            col_widths = [8 * mm, 10 * mm, 45 * mm] + [7 * mm] * grid['num_days']
            t = Table(table_data, colWidths=col_widths, repeatRows=1)

            style_commands = [
                ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1e3a5f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (2, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#475569')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.HexColor('#f8fafc'), rl_colors.white]),
            ]

            # Color code cells
            for row_idx in range(1, len(table_data)):
                for col_idx in range(3, len(table_data[row_idx])):
                    shift = table_data[row_idx][col_idx]
                    if shift in SHIFT_PDF_COLORS:
                        style_commands.append(
                            ('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx),
                             SHIFT_PDF_COLORS[shift])
                        )
                        style_commands.append(
                            ('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx),
                             rl_colors.white)
                        )

            t.setStyle(TableStyle(style_commands))
            elements.append(t)
            elements.append(Spacer(1, 5 * mm))

    doc.build(elements)
    output.seek(0)
    return output


def export_audit_excel(filters=None):
    """Export audit logs to Excel."""
    query = AuditLog.query.order_by(AuditLog.timestamp.desc())

    if filters:
        if filters.get('group'):
            group_num = int(filters['group'])
            worker_ids = [w.id for w in Worker.query.filter_by(group_number=group_num).all()]
            query = query.filter(AuditLog.target_worker_id.in_(worker_ids))
        if filters.get('start_date'):
            query = query.filter(AuditLog.timestamp >= filters['start_date'])
        if filters.get('end_date'):
            query = query.filter(AuditLog.timestamp <= filters['end_date'])
        if filters.get('action'):
            query = query.filter_by(action=filters['action'])

    logs = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Registro de Auditoría'

    headers = ['Fecha/Hora', 'Usuario', 'Acción', 'Trabajador', 'Fecha Afectada',
               'Valor Anterior', 'Valor Nuevo', 'Detalles']

    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, log in enumerate(logs, 2):
        ws.cell(row=row_idx, column=1, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_idx, column=2, value=log.user.username if log.user else '')
        ws.cell(row=row_idx, column=3, value=AuditLog.ACTION_LABELS.get(log.action, log.action))
        ws.cell(row=row_idx, column=4,
                value=log.target_worker.full_name if log.target_worker else '')
        ws.cell(row=row_idx, column=5,
                value=log.target_date.isoformat() if log.target_date else '')
        ws.cell(row=row_idx, column=6, value=log.old_value or '')
        ws.cell(row=row_idx, column=7, value=log.new_value or '')
        ws.cell(row=row_idx, column=8, value=log.details or '')

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

